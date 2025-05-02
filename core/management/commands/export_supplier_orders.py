from django.core.management.base import BaseCommand
import openpyxl
from openpyxl.utils import get_column_letter
from core.models.supplier_order import SupplierOrder

class Command(BaseCommand):
    help = "Export Supplier Orders into an Excel file."

    def handle(self, *args, **kwargs):
        filename = 'supplier_orders_export.xlsx'
        self.stdout.write(f"[INFO] Exporting supplier orders to {filename}...")

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Supplier Orders"

        # Define columns and fields
        columns = [
            "Date", "Book No.", "Order No.", "Tax Invoice", "Supplier",
            "PC", "Stone", "H/NH", "Color", "Shape", "Size",
            "Carats", "Currency", "Price per Unit", "Unit",
            "Total THB", "Weight per Piece", "Price $/ct",
            "Price $ per Piece", "Total USD", "Rate Avg 2019",
            "Remarks", "Credit Term", "Target Size"
        ]
        fields = [
            "date", "book_no", "order_no", "tax_invoice", "supplier",
            "number", "stone", "heating", "color", "shape", "size",
            "carats", "currency", "price_cur_per_unit", "unit",
            "total_thb", "weight_per_piece", "price_usd_per_ct",
            "price_usd_per_piece", "total_usd", "rate_avg_2019",
            "remarks", "credit_term", "target_size"
        ]

        # Write header
        for col_num, column_title in enumerate(columns, 1):
            worksheet.cell(row=1, column=col_num, value=column_title)

        # Write rows
        orders = SupplierOrder.objects.all().order_by('-date')
        for row_num, order in enumerate(orders, 2):
            row = []
            for field in fields:
                value = getattr(order, field)
                if hasattr(value, 'strftime'):
                    value = value.strftime('%Y-%m-%d')
                elif value is None:
                    value = ''
                row.append(value)

            for col_num, cell_value in enumerate(row, 1):
                worksheet.cell(row=row_num, column=col_num, value=cell_value)

        # Adjust column widths
        for i, _ in enumerate(columns, 1):
            worksheet.column_dimensions[get_column_letter(i)].width = 15

        # Save to file
        workbook.save(filename)

        self.stdout.write(self.style.SUCCESS(f"[DONE] Exported {orders.count()} supplier orders to '{filename}'"))
