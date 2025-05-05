import openpyxl
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def export_orders_to_excel(queryset, filename="orders.xlsx", columns=None, fields=None):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Orders"

    # Write headers
    for col_num, column_title in enumerate(columns, 1):
        worksheet.cell(row=1, column=col_num, value=column_title)

    # Write data
    for row_num, obj in enumerate(queryset, 2):
        row = []
        for field in fields:
            value = getattr(obj, field)
            # Special format for dates and decimals
            if hasattr(value, 'strftime'):
                value = value.strftime('%Y-%m-%d')
            elif value is None:
                value = ''
            row.append(value)

        for col_num, cell_value in enumerate(row, 1):
            worksheet.cell(row=row_num, column=col_num, value=cell_value)

    # Adjust column widths (optional)
    for i, _ in enumerate(columns, 1):
        worksheet.column_dimensions[get_column_letter(i)].width = 15

    # Prepare response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    workbook.save(response)

    return response
